"""
Property Comparison Tool - Backend API
========================================
Purpose: Import real estate deals from Excel models, standardize metrics, and generate comparisons
Supports: Multifamily, Mixed-Use, Hotel, SFR, House Flipping models
Created: November 4, 2025
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends
from fastapi.responses import FileResponse
from pydantic import BaseModel, UUID4, condecimal, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date
from decimal import Decimal
import uuid
import openpyxl
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor, execute_values
import json
from pathlib import Path

# =====================================================
# PYDANTIC MODELS (Data Validation)
# =====================================================

class ComparisonSetCreate(BaseModel):
    """Create a new comparison set"""
    comparison_name: str = Field(..., min_length=1, max_length=255)
    comparison_description: Optional[str] = None
    primary_metric: str = "levered_irr"

class ComparisonSetResponse(BaseModel):
    """Response model for comparison set"""
    comparison_id: UUID4
    comparison_name: str
    comparison_description: Optional[str]
    status: str
    created_at: datetime
    deal_count: int = 0

class PropertyDealImport(BaseModel):
    """Model for importing a property deal"""
    comparison_id: UUID4
    property_name: str
    property_address: Optional[str] = None
    property_type: str  # Multifamily, Mixed-Use, Hotel, SFR, House Flipping
    property_subtype: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    file_path: str  # Path to the Excel model

class ComparisonMetrics(BaseModel):
    """Standardized metrics for property comparison"""
    # Returns
    levered_irr: Optional[condecimal(max_digits=8, decimal_places=4)] = None
    unlevered_irr: Optional[condecimal(max_digits=8, decimal_places=4)] = None
    equity_multiple: Optional[condecimal(max_digits=8, decimal_places=4)] = None
    cash_on_cash_y1: Optional[condecimal(max_digits=8, decimal_places=4)] = None
    
    # Valuation
    entry_cap_rate: Optional[condecimal(max_digits=6, decimal_places=4)] = None
    exit_cap_rate: Optional[condecimal(max_digits=6, decimal_places=4)] = None
    
    # Operations
    noi_year1: Optional[Decimal] = None
    noi_stabilized: Optional[Decimal] = None
    noi_margin: Optional[condecimal(max_digits=6, decimal_places=4)] = None
    
    # Debt
    dscr_year1: Optional[condecimal(max_digits=8, decimal_places=4)] = None
    ltv: Optional[condecimal(max_digits=6, decimal_places=4)] = None
    
    # Investment Size
    purchase_price: Optional[Decimal] = None
    equity_required: Optional[Decimal] = None

class DealComparisonResponse(BaseModel):
    """Response with all deals in comparison"""
    deal_id: UUID4
    property_name: str
    property_type: str
    city: Optional[str]
    state: Optional[str]
    metrics: ComparisonMetrics
    overall_score: Optional[Decimal]
    overall_rank: Optional[int]

class ScoringCriteria(BaseModel):
    """Scoring criteria for deals"""
    category: str  # Returns, Risk, Location, Operations
    criteria_name: str
    weight: condecimal(max_digits=5, decimal_places=4)
    metric_field: str
    excellent_threshold: Optional[Decimal] = None
    good_threshold: Optional[Decimal] = None
    acceptable_threshold: Optional[Decimal] = None
    poor_threshold: Optional[Decimal] = None

# =====================================================
# FASTAPI APP
# =====================================================

app = FastAPI(
    title="Property Comparison Tool API",
    description="Compare real estate deals from multiple Excel models",
    version="1.0.0"
)

# =====================================================
# DATABASE CONNECTION
# =====================================================

def get_db_connection():
    """Get PostgreSQL database connection"""
    try:
        conn = psycopg2.connect(
            host="localhost",
            database="portfolio_dashboard",
            user="postgres",
            password="your_password",  # Update with actual password
            cursor_factory=RealDictCursor
        )
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database connection failed: {str(e)}")

# =====================================================
# EXCEL IMPORT FUNCTIONS (Model-Specific)
# =====================================================

class ExcelModelImporter:
    """Import deals from various Excel models"""
    
    @staticmethod
    def import_multifamily_model(file_path: str) -> Dict[str, Any]:
        """Import from Multifamily Model"""
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Navigate to specific sheets and cells based on model structure
            # Executive Summary or Returns Analysis sheet
            if 'Executive Summary' in wb.sheetnames:
                ws = wb['Executive Summary']
            elif 'Returns Analysis' in wb.sheetnames:
                ws = wb['Returns Analysis']
            else:
                raise ValueError("Could not find expected sheets in Multifamily Model")
            
            # Extract metrics (adjust cell references based on actual model)
            metrics = {
                'levered_irr': ws['B10'].value if ws['B10'].value else None,  # Example cell
                'unlevered_irr': ws['B11'].value if ws['B11'].value else None,
                'equity_multiple': ws['B12'].value if ws['B12'].value else None,
                'cash_on_cash_y1': ws['B13'].value if ws['B13'].value else None,
                'entry_cap_rate': ws['B15'].value if ws['B15'].value else None,
                'exit_cap_rate': ws['B16'].value if ws['B16'].value else None,
                'noi_year1': ws['B20'].value if ws['B20'].value else None,
                'noi_stabilized': ws['B21'].value if ws['B21'].value else None,
                'dscr_year1': ws['B25'].value if ws['B25'].value else None,
                'equity_required': ws['B30'].value if ws['B30'].value else None,
                'purchase_price': ws['B31'].value if ws['B31'].value else None,
            }
            
            # Get property details from Inputs sheet
            inputs_ws = wb['Inputs'] if 'Inputs' in wb.sheetnames else None
            if inputs_ws:
                metrics['property_name'] = inputs_ws['B5'].value or "Unnamed Property"
                metrics['total_sf'] = inputs_ws['B10'].value
                metrics['units'] = inputs_ws['B11'].value
                metrics['city'] = inputs_ws['B6'].value
                metrics['state'] = inputs_ws['B7'].value
            
            return metrics
            
        except Exception as e:
            raise ValueError(f"Error importing Multifamily model: {str(e)}")
    
    @staticmethod
    def import_mixed_use_model(file_path: str) -> Dict[str, Any]:
        """Import from Mixed-Use Model"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['Executive Summary'] if 'Executive Summary' in wb.sheetnames else wb['Returns Analysis']
            
            metrics = {
                'levered_irr': ws['B12'].value,
                'equity_multiple': ws['B13'].value,
                'cash_on_cash_y1': ws['B14'].value,
                'entry_cap_rate': ws['B16'].value,
                'exit_cap_rate': ws['B17'].value,
                'noi_stabilized': ws['B20'].value,
                'dscr_year1': ws['B25'].value,
                'equity_required': ws['B30'].value,
                'total_project_cost': ws['B31'].value,
            }
            
            # Get property info
            inputs_ws = wb['Inputs']
            metrics['property_name'] = inputs_ws['B5'].value or "Unnamed Mixed-Use"
            metrics['total_sf'] = inputs_ws['B12'].value
            metrics['city'] = inputs_ws['B7'].value
            metrics['state'] = inputs_ws['B8'].value
            
            return metrics
            
        except Exception as e:
            raise ValueError(f"Error importing Mixed-Use model: {str(e)}")
    
    @staticmethod
    def import_hotel_model(file_path: str) -> Dict[str, Any]:
        """Import from Hotel Model"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['Executive Summary']
            
            metrics = {
                'levered_irr': ws['B10'].value,
                'equity_multiple': ws['B11'].value,
                'cash_on_cash_y1': ws['B12'].value,
                'entry_cap_rate': ws['B14'].value,
                'exit_cap_rate': ws['B15'].value,
                'noi_year1': ws['B18'].value,
                'dscr_year1': ws['B22'].value,
                'equity_required': ws['B26'].value,
            }
            
            inputs_ws = wb['Inputs']
            metrics['property_name'] = inputs_ws['B5'].value or "Unnamed Hotel"
            metrics['units'] = inputs_ws['B10'].value  # Number of rooms
            metrics['city'] = inputs_ws['B6'].value
            metrics['state'] = inputs_ws['B7'].value
            
            return metrics
            
        except Exception as e:
            raise ValueError(f"Error importing Hotel model: {str(e)}")
    
    @staticmethod
    def import_sfr_model(file_path: str) -> Dict[str, Any]:
        """Import from Single-Family Rental Model"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['Executive Summary'] if 'Executive Summary' in wb.sheetnames else wb['Returns']
            
            metrics = {
                'levered_irr': ws['B8'].value,
                'equity_multiple': ws['B9'].value,
                'cash_on_cash_y1': ws['B10'].value,
                'entry_cap_rate': ws['B12'].value,
                'noi_year1': ws['B15'].value,
                'dscr_year1': ws['B18'].value,
                'equity_required': ws['B22'].value,
                'purchase_price': ws['B21'].value,
            }
            
            inputs_ws = wb['Inputs']
            metrics['property_name'] = inputs_ws['B5'].value or "Unnamed SFR"
            metrics['total_sf'] = inputs_ws['B9'].value
            metrics['city'] = inputs_ws['B6'].value
            metrics['state'] = inputs_ws['B7'].value
            
            return metrics
            
        except Exception as e:
            raise ValueError(f"Error importing SFR model: {str(e)}")
    
    @staticmethod
    def import_house_flipping_model(file_path: str) -> Dict[str, Any]:
        """Import from House Flipping Model"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb['Executive Summary']
            
            metrics = {
                'levered_irr': ws['B10'].value,
                'equity_multiple': ws['B11'].value,
                'cash_on_cash_y1': ws['B12'].value,
                'purchase_price': ws['B15'].value,
                'equity_required': ws['B20'].value,
                'hold_period_months': ws['B18'].value,
            }
            
            inputs_ws = wb['Inputs']
            metrics['property_name'] = inputs_ws['B5'].value or "Unnamed Flip"
            metrics['total_sf'] = inputs_ws['B9'].value
            metrics['city'] = inputs_ws['B6'].value
            metrics['state'] = inputs_ws['B7'].value
            
            return metrics
            
        except Exception as e:
            raise ValueError(f"Error importing House Flipping model: {str(e)}")
    
    @staticmethod
    def import_model(file_path: str, model_type: str) -> Dict[str, Any]:
        """Route to appropriate importer based on model type"""
        importers = {
            'Multifamily': ExcelModelImporter.import_multifamily_model,
            'Mixed-Use': ExcelModelImporter.import_mixed_use_model,
            'Hotel': ExcelModelImporter.import_hotel_model,
            'SFR': ExcelModelImporter.import_sfr_model,
            'House Flipping': ExcelModelImporter.import_house_flipping_model,
        }
        
        if model_type not in importers:
            raise ValueError(f"Unsupported model type: {model_type}")
        
        return importers[model_type](file_path)

# =====================================================
# SCORING ENGINE
# =====================================================

class ScoringEngine:
    """Score and rank deals based on weighted criteria"""
    
    @staticmethod
    def calculate_metric_score(value: float, excellent: float, good: float, 
                               acceptable: float, poor: float, 
                               scoring_method: str = 'linear') -> float:
        """
        Calculate 0-100 score based on thresholds
        
        Args:
            value: Actual metric value
            excellent: Threshold for 100 score
            good: Threshold for 75 score
            acceptable: Threshold for 50 score
            poor: Threshold for 25 score
            scoring_method: 'linear' or 'inverse' (lower is better)
        """
        if value is None:
            return 0.0
        
        if scoring_method == 'inverse':
            # For metrics where lower is better (e.g., LTV)
            if value <= excellent:
                return 100.0
            elif value <= good:
                return 75.0 + 25.0 * (good - value) / (good - excellent)
            elif value <= acceptable:
                return 50.0 + 25.0 * (acceptable - value) / (acceptable - good)
            elif value <= poor:
                return 25.0 + 25.0 * (poor - value) / (poor - acceptable)
            else:
                return max(0.0, 25.0 * (1 - (value - poor) / poor))
        
        else:  # linear scoring (higher is better)
            if value >= excellent:
                return 100.0
            elif value >= good:
                return 75.0 + 25.0 * (value - good) / (excellent - good)
            elif value >= acceptable:
                return 50.0 + 25.0 * (value - acceptable) / (good - acceptable)
            elif value >= poor:
                return 25.0 + 25.0 * (value - poor) / (acceptable - poor)
            else:
                return max(0.0, 25.0 * value / poor) if poor > 0 else 0.0
    
    @staticmethod
    def score_deal(deal_id: str, comparison_id: str, conn):
        """Score a deal against all active criteria"""
        cur = conn.cursor()
        
        # Get the deal's metrics
        cur.execute("""
            SELECT * FROM comparison_metrics 
            WHERE deal_id = %s
        """, (deal_id,))
        metrics = cur.fetchone()
        
        if not metrics:
            return None
        
        # Get active scoring criteria
        cur.execute("""
            SELECT * FROM scoring_criteria 
            WHERE comparison_id = %s AND is_active = TRUE
        """, (comparison_id,))
        criteria = cur.fetchall()
        
        total_weighted_score = 0.0
        category_scores = {}
        
        for criterion in criteria:
            metric_value = metrics.get(criterion['metric_field'])
            
            if metric_value is not None:
                # Calculate individual score
                score = ScoringEngine.calculate_metric_score(
                    float(metric_value),
                    float(criterion['excellent_threshold']) if criterion['excellent_threshold'] else 100,
                    float(criterion['good_threshold']) if criterion['good_threshold'] else 75,
                    float(criterion['acceptable_threshold']) if criterion['acceptable_threshold'] else 50,
                    float(criterion['poor_threshold']) if criterion['poor_threshold'] else 25,
                    criterion['scoring_method']
                )
                
                weighted_score = score * float(criterion['weight'])
                total_weighted_score += weighted_score
                
                # Track category scores
                category = criterion['category']
                if category not in category_scores:
                    category_scores[category] = 0.0
                category_scores[category] += weighted_score
                
                # Insert individual score
                cur.execute("""
                    INSERT INTO deal_scores 
                    (deal_id, criteria_id, raw_value, normalized_score, weighted_score)
                    VALUES (%s, %s, %s, %s, %s)
                """, (deal_id, criterion['criteria_id'], metric_value, score, weighted_score))
        
        # Update overall score in comparison_metrics
        cur.execute("""
            UPDATE comparison_metrics 
            SET overall_score = %s,
                returns_score = %s,
                risk_score = %s,
                location_score = %s,
                operational_score = %s
            WHERE deal_id = %s
        """, (
            total_weighted_score,
            category_scores.get('Returns', 0),
            category_scores.get('Risk', 0),
            category_scores.get('Location', 0),
            category_scores.get('Operations', 0),
            deal_id
        ))
        
        conn.commit()
        cur.close()
        
        return {
            'overall_score': total_weighted_score,
            'category_scores': category_scores
        }
    
    @staticmethod
    def rank_deals(comparison_id: str, conn):
        """Rank all deals in a comparison by overall score"""
        cur = conn.cursor()
        
        # Rank by overall score
        cur.execute("""
            WITH ranked AS (
                SELECT 
                    pd.deal_id,
                    cm.overall_score,
                    ROW_NUMBER() OVER (ORDER BY cm.overall_score DESC) as rank
                FROM property_deals pd
                JOIN comparison_metrics cm ON pd.deal_id = cm.deal_id
                WHERE pd.comparison_id = %s
                  AND pd.deal_status = 'Active'
            )
            UPDATE property_deals
            SET overall_rank = ranked.rank
            FROM ranked
            WHERE property_deals.deal_id = ranked.deal_id
        """, (comparison_id,))
        
        # Rank by risk-adjusted return
        cur.execute("""
            WITH risk_ranked AS (
                SELECT 
                    pd.deal_id,
                    (cm.levered_irr * (cm.risk_score / 100.0)) as risk_adj_irr,
                    ROW_NUMBER() OVER (ORDER BY (cm.levered_irr * (cm.risk_score / 100.0)) DESC) as rank
                FROM property_deals pd
                JOIN comparison_metrics cm ON pd.deal_id = cm.deal_id
                WHERE pd.comparison_id = %s
                  AND pd.deal_status = 'Active'
            )
            UPDATE property_deals
            SET risk_adjusted_rank = risk_ranked.rank
            FROM risk_ranked
            WHERE property_deals.deal_id = risk_ranked.deal_id
        """, (comparison_id,))
        
        conn.commit()
        cur.close()

# =====================================================
# API ENDPOINTS
# =====================================================

@app.post("/api/comparisons", response_model=ComparisonSetResponse)
async def create_comparison(comparison: ComparisonSetCreate):
    """Create a new comparison set"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        comparison_id = str(uuid.uuid4())
        
        cur.execute("""
            INSERT INTO comparison_sets 
            (comparison_id, comparison_name, comparison_description, primary_metric)
            VALUES (%s, %s, %s, %s)
            RETURNING comparison_id, comparison_name, comparison_description, status, created_at
        """, (comparison_id, comparison.comparison_name, comparison.comparison_description, 
              comparison.primary_metric))
        
        result = cur.fetchone()
        conn.commit()
        
        # Insert default scoring criteria
        insert_default_criteria(comparison_id, conn)
        
        return ComparisonSetResponse(
            comparison_id=result['comparison_id'],
            comparison_name=result['comparison_name'],
            comparison_description=result['comparison_description'],
            status=result['status'],
            created_at=result['created_at'],
            deal_count=0
        )
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.post("/api/comparisons/{comparison_id}/deals/import")
async def import_deal(comparison_id: str, file: UploadFile = File(...), 
                     property_type: str = "Multifamily"):
    """Import a deal from an Excel model"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Save uploaded file temporarily
        temp_path = f"/tmp/{file.filename}"
        with open(temp_path, "wb") as f:
            f.write(await file.read())
        
        # Import metrics from Excel
        metrics = ExcelModelImporter.import_model(temp_path, property_type)
        
        # Create deal record
        deal_id = str(uuid.uuid4())
        cur.execute("""
            INSERT INTO property_deals 
            (deal_id, comparison_id, property_name, property_type, city, state, 
             source_model_type, source_file_path, total_sf, units)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING deal_id
        """, (deal_id, comparison_id, metrics.get('property_name'), property_type,
              metrics.get('city'), metrics.get('state'), property_type, temp_path,
              metrics.get('total_sf'), metrics.get('units')))
        
        # Insert metrics
        cur.execute("""
            INSERT INTO comparison_metrics 
            (deal_id, levered_irr, unlevered_irr, equity_multiple, cash_on_cash_y1,
             entry_cap_rate, exit_cap_rate, noi_year1, noi_stabilized, dscr_year1,
             equity_required, purchase_price)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (deal_id, metrics.get('levered_irr'), metrics.get('unlevered_irr'),
              metrics.get('equity_multiple'), metrics.get('cash_on_cash_y1'),
              metrics.get('entry_cap_rate'), metrics.get('exit_cap_rate'),
              metrics.get('noi_year1'), metrics.get('noi_stabilized'),
              metrics.get('dscr_year1'), metrics.get('equity_required'),
              metrics.get('purchase_price')))
        
        conn.commit()
        
        # Score the deal
        ScoringEngine.score_deal(deal_id, comparison_id, conn)
        ScoringEngine.rank_deals(comparison_id, conn)
        
        return {"deal_id": deal_id, "status": "imported", "metrics": metrics}
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.get("/api/comparisons/{comparison_id}/deals")
async def get_comparison_deals(comparison_id: str):
    """Get all deals in a comparison with scores and rankings"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT 
                pd.*,
                cm.*
            FROM property_deals pd
            JOIN comparison_metrics cm ON pd.deal_id = cm.deal_id
            WHERE pd.comparison_id = %s
            ORDER BY cm.overall_score DESC
        """, (comparison_id,))
        
        deals = cur.fetchall()
        return {"comparison_id": comparison_id, "deals": deals}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.get("/api/comparisons/{comparison_id}/heatmap")
async def get_heatmap_data(comparison_id: str):
    """Get data formatted for heatmap visualization"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Get all deals with normalized scores by metric
        cur.execute("""
            SELECT 
                pd.property_name,
                pd.property_type,
                cm.levered_irr,
                cm.equity_multiple,
                cm.cash_on_cash_y1,
                cm.dscr_year1,
                cm.entry_cap_rate,
                cm.noi_margin,
                cm.overall_score,
                cm.returns_score,
                cm.risk_score,
                pd.overall_rank
            FROM property_deals pd
            JOIN comparison_metrics cm ON pd.deal_id = cm.deal_id
            WHERE pd.comparison_id = %s
            ORDER BY cm.overall_score DESC
        """, (comparison_id,))
        
        deals = cur.fetchall()
        
        # Normalize metrics to 0-100 scale for heatmap
        metrics_list = ['levered_irr', 'equity_multiple', 'cash_on_cash_y1', 
                       'dscr_year1', 'entry_cap_rate', 'noi_margin']
        
        heatmap_data = []
        for deal in deals:
            row = {
                'property_name': deal['property_name'],
                'property_type': deal['property_type'],
                'overall_rank': deal['overall_rank'],
                'metrics': {}
            }
            
            for metric in metrics_list:
                value = deal[metric]
                if value is not None:
                    # Normalize based on min/max in dataset
                    row['metrics'][metric] = float(value)
            
            heatmap_data.append(row)
        
        return {"heatmap_data": heatmap_data}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# =====================================================
# HELPER FUNCTIONS
# =====================================================

def insert_default_criteria(comparison_id: str, conn):
    """Insert default scoring criteria for a new comparison"""
    cur = conn.cursor()
    
    default_criteria = [
        # RETURNS (40% total weight)
        ('Returns', 'Levered IRR', 0.25, 'levered_irr', 'linear', 0.25, 0.20, 0.15, 0.10),
        ('Returns', 'Equity Multiple', 0.10, 'equity_multiple', 'linear', 2.50, 2.00, 1.60, 1.30),
        ('Returns', 'Cash-on-Cash Y1', 0.05, 'cash_on_cash_y1', 'linear', 0.12, 0.08, 0.05, 0.03),
        
        # RISK (30% total weight)
        ('Risk', 'DSCR Y1', 0.15, 'dscr_year1', 'linear', 1.50, 1.35, 1.25, 1.15),
        ('Risk', 'LTV', 0.10, 'ltv', 'inverse', 0.65, 0.70, 0.75, 0.80),
        ('Risk', 'Cap Rate Spread', 0.05, 'cap_rate_spread', 'linear', 150, 100, 50, 0),
        
        # OPERATIONS (30% total weight)
        ('Operations', 'NOI Margin', 0.15, 'noi_margin', 'linear', 0.65, 0.60, 0.55, 0.50),
        ('Operations', 'Stabilized Occupancy', 0.15, 'occupancy_stabilized', 'linear', 0.95, 0.92, 0.88, 0.85),
    ]
    
    for criteria in default_criteria:
        cur.execute("""
            INSERT INTO scoring_criteria 
            (comparison_id, category, criteria_name, weight, metric_field, scoring_method,
             excellent_threshold, good_threshold, acceptable_threshold, poor_threshold)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (comparison_id, *criteria))
    
    conn.commit()
    cur.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
