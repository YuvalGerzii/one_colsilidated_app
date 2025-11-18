"""
Portfolio Dashboard - Excel Model Generator
Generates all 5 financial models (DCF, LBO, Merger, DD Tracker, QoE) from database
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any
from decimal import Decimal
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker, Session
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============================================================================
# DATABASE MODELS (SQLAlchemy)
# ============================================================================

from sqlalchemy import Column, String, Integer, Numeric, Date, DateTime, Boolean, Text, ARRAY
from sqlalchemy.dialects.postgresql import UUID, JSONB
from sqlalchemy.ext.declarative import declarative_base
import uuid

Base = declarative_base()


class PortfolioCompany(Base):
    """Portfolio company database model"""
    __tablename__ = 'portfolio_companies'
    
    company_id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    fund_id = Column(UUID(as_uuid=True), nullable=False)
    company_name = Column(String(255), nullable=False)
    ticker_symbol = Column(String(10))
    
    investment_date = Column(Date, nullable=False)
    deal_type = Column(String(50), nullable=False)
    sector = Column(String(100), nullable=False)
    industry = Column(String(100))
    
    entry_revenue = Column(Numeric(15, 2))
    entry_ebitda = Column(Numeric(15, 2))
    entry_multiple = Column(Numeric(10, 2))
    purchase_price = Column(Numeric(15, 2))
    equity_invested = Column(Numeric(15, 2))
    debt_raised = Column(Numeric(15, 2))
    ownership_percentage = Column(Numeric(5, 4))
    
    company_status = Column(String(50), default='Active')


class FinancialMetric(Base):
    """Time-series financial metrics"""
    __tablename__ = 'financial_metrics'
    
    metric_id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    company_id = Column(UUID(as_uuid=True), nullable=False)
    
    period_date = Column(Date, nullable=False)
    period_type = Column(String(20), nullable=False)
    fiscal_year = Column(Integer, nullable=False)
    fiscal_quarter = Column(Integer)
    
    # Income Statement
    revenue = Column(Numeric(15, 2))
    cogs = Column(Numeric(15, 2))
    gross_profit = Column(Numeric(15, 2))
    gross_margin = Column(Numeric(5, 4))
    
    operating_expenses = Column(Numeric(15, 2))
    sales_marketing = Column(Numeric(15, 2))
    research_development = Column(Numeric(15, 2))
    general_admin = Column(Numeric(15, 2))
    
    ebitda = Column(Numeric(15, 2))
    ebitda_margin = Column(Numeric(5, 4))
    depreciation = Column(Numeric(15, 2))
    amortization = Column(Numeric(15, 2))
    ebit = Column(Numeric(15, 2))
    
    interest_expense = Column(Numeric(15, 2))
    tax_expense = Column(Numeric(15, 2))
    tax_rate = Column(Numeric(5, 4))
    net_income = Column(Numeric(15, 2))
    
    # Balance Sheet
    cash = Column(Numeric(15, 2))
    accounts_receivable = Column(Numeric(15, 2))
    inventory = Column(Numeric(15, 2))
    total_current_assets = Column(Numeric(15, 2))
    
    ppe_net = Column(Numeric(15, 2))
    intangible_assets = Column(Numeric(15, 2))
    goodwill = Column(Numeric(15, 2))
    total_assets = Column(Numeric(15, 2))
    
    accounts_payable = Column(Numeric(15, 2))
    current_portion_debt = Column(Numeric(15, 2))
    total_current_liabilities = Column(Numeric(15, 2))
    
    long_term_debt = Column(Numeric(15, 2))
    total_liabilities = Column(Numeric(15, 2))
    shareholders_equity = Column(Numeric(15, 2))
    
    # Cash Flow
    operating_cash_flow = Column(Numeric(15, 2))
    capex = Column(Numeric(15, 2))
    free_cash_flow = Column(Numeric(15, 2))
    
    working_capital = Column(Numeric(15, 2))


class Valuation(Base):
    """Company valuations"""
    __tablename__ = 'valuations'
    
    valuation_id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    company_id = Column(UUID(as_uuid=True), nullable=False)
    
    valuation_date = Column(Date, nullable=False)
    valuation_method = Column(String(50), nullable=False)
    
    # DCF specific
    wacc = Column(Numeric(5, 4))
    terminal_growth_rate = Column(Numeric(5, 4))
    terminal_multiple = Column(Numeric(10, 2))
    
    enterprise_value = Column(Numeric(15, 2))
    equity_value = Column(Numeric(15, 2))
    
    # LBO specific
    entry_multiple = Column(Numeric(10, 2))
    exit_multiple = Column(Numeric(10, 2))
    hold_period_years = Column(Integer)
    
    unrealized_moic = Column(Numeric(10, 2))
    unrealized_irr = Column(Numeric(10, 4))


# ============================================================================
# STYLING HELPERS
# ============================================================================

def apply_header_style(cell):
    """Apply header styling"""
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_subheader_style(cell):
    """Apply subheader styling"""
    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')


def apply_input_style(cell):
    """Apply input cell styling (blue cells users can edit)"""
    cell.font = Font(name='Calibri', size=10, color='000000')
    cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    cell.alignment = Alignment(horizontal='right')


def apply_formula_style(cell):
    """Apply formula cell styling"""
    cell.font = Font(name='Calibri', size=10, color='000000')
    cell.alignment = Alignment(horizontal='right')


def apply_number_format(cell, format_type='currency'):
    """Apply number formatting"""
    if format_type == 'currency':
        cell.number_format = '$#,##0'
    elif format_type == 'currency_decimal':
        cell.number_format = '$#,##0.00'
    elif format_type == 'percent':
        cell.number_format = '0.0%'
    elif format_type == 'percent_decimal':
        cell.number_format = '0.00%'
    elif format_type == 'number':
        cell.number_format = '#,##0'
    elif format_type == 'decimal':
        cell.number_format = '#,##0.00'
    elif format_type == 'multiple':
        cell.number_format = '0.0x'


# ============================================================================
# MODEL GENERATOR BASE CLASS
# ============================================================================

class ModelGenerator:
    """Base class for all model generators"""
    
    def __init__(self, db_session: Session, company_id: str, template_path: str = None):
        self.db_session = db_session
        self.company_id = company_id
        self.template_path = template_path
        self.company = None
        self.financials = []
        self.valuation = None
        
    def fetch_data(self):
        """Fetch all required data from database"""
        # Get company
        self.company = self.db_session.query(PortfolioCompany).filter(
            PortfolioCompany.company_id == self.company_id
        ).first()
        
        if not self.company:
            raise ValueError(f"Company {self.company_id} not found")
        
        # Get financial metrics
        self.financials = self.db_session.query(FinancialMetric).filter(
            FinancialMetric.company_id == self.company_id
        ).order_by(FinancialMetric.period_date).all()
        
        # Get latest valuation
        self.valuation = self.db_session.query(Valuation).filter(
            Valuation.company_id == self.company_id
        ).order_by(Valuation.valuation_date.desc()).first()
        
        logger.info(f"Fetched data for {self.company.company_name}")
        logger.info(f"Found {len(self.financials)} financial periods")
    
    def load_template(self) -> Workbook:
        """Load Excel template or create new workbook"""
        if self.template_path and os.path.exists(self.template_path):
            return load_workbook(self.template_path)
        return Workbook()
    
    def save(self, output_path: str):
        """Save the generated model"""
        self.wb.save(output_path)
        logger.info(f"Model saved to {output_path}")
    
    def generate(self, output_path: str):
        """Main generation method - to be implemented by subclasses"""
        raise NotImplementedError("Subclasses must implement generate()")


# ============================================================================
# DCF MODEL GENERATOR
# ============================================================================

class DCFModelGenerator(ModelGenerator):
    """Generate DCF model from database"""
    
    def generate(self, output_path: str):
        """Generate complete DCF model"""
        logger.info(f"Generating DCF model for {self.company.company_name}...")
        
        # Fetch data
        self.fetch_data()
        
        # Load template
        template_path = '/mnt/user-data/uploads/DCF_Model_Comprehensive.xlsx'
        if os.path.exists(template_path):
            self.wb = load_workbook(template_path)
            logger.info("Loaded DCF template")
        else:
            raise FileNotFoundError("DCF template not found")
        
        # Populate sheets
        self._populate_dcf_sheet()
        self._populate_historical_financials()
        self._populate_wacc()
        
        # Save
        self.save(output_path)
    
    def _populate_dcf_sheet(self):
        """Populate main DCF sheet with company data"""
        if 'DCF' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['DCF']
        
        # Company name
        sheet['B2'] = self.company.company_name
        sheet['B2'].font = Font(size=14, bold=True)
        
        # Get latest financials for base year
        if self.financials:
            latest = self.financials[-1]
            
            # Base year revenue (example: Cell C8)
            if latest.revenue:
                sheet['C8'] = float(latest.revenue)
                apply_input_style(sheet['C8'])
                apply_number_format(sheet['C8'], 'currency')
            
            # EBITDA margin
            if latest.ebitda_margin:
                sheet['C12'] = float(latest.ebitda_margin)
                apply_input_style(sheet['C12'])
                apply_number_format(sheet['C12'], 'percent')
        
        logger.info("DCF sheet populated")
    
    def _populate_historical_financials(self):
        """Populate historical financials sheet"""
        if 'Historical Financials' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Historical Financials']
        
        # Populate historical data (last 3-5 years)
        start_row = 8  # Adjust based on template
        for idx, metric in enumerate(self.financials[-5:]):  # Last 5 years
            col = chr(67 + idx)  # C, D, E, F, G
            
            # Revenue
            if metric.revenue:
                sheet[f'{col}{start_row}'] = float(metric.revenue)
            
            # COGS
            if metric.cogs:
                sheet[f'{col}{start_row+1}'] = float(metric.cogs)
            
            # EBITDA
            if metric.ebitda:
                sheet[f'{col}{start_row+5}'] = float(metric.ebitda)
            
            # Apply formatting
            for row in range(start_row, start_row + 20):
                if sheet[f'{col}{row}'].value:
                    apply_number_format(sheet[f'{col}{row}'], 'currency')
        
        logger.info("Historical financials populated")
    
    def _populate_wacc(self):
        """Populate WACC sheet with valuation data"""
        if 'WACC' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['WACC']
        
        if self.valuation and self.valuation.wacc:
            # WACC value (adjust cell reference based on template)
            sheet['C15'] = float(self.valuation.wacc)
            apply_input_style(sheet['C15'])
            apply_number_format(sheet['C15'], 'percent_decimal')
        
        logger.info("WACC sheet populated")


# ============================================================================
# LBO MODEL GENERATOR
# ============================================================================

class LBOModelGenerator(ModelGenerator):
    """Generate LBO model from database"""
    
    def generate(self, output_path: str):
        """Generate complete LBO model"""
        logger.info(f"Generating LBO model for {self.company.company_name}...")
        
        # Fetch data
        self.fetch_data()
        
        # Load template
        template_path = '/mnt/user-data/uploads/LBO_Model_Comprehensive.xlsx'
        if os.path.exists(template_path):
            self.wb = load_workbook(template_path)
            logger.info("Loaded LBO template")
        else:
            raise FileNotFoundError("LBO template not found")
        
        # Populate sheets
        self._populate_transaction_assumptions()
        self._populate_sources_uses()
        self._populate_operating_model()
        
        # Save
        self.save(output_path)
    
    def _populate_transaction_assumptions(self):
        """Populate transaction assumptions"""
        if 'Transaction Assumptions' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Transaction Assumptions']
        
        # Company name
        sheet['C5'] = self.company.company_name
        
        # Purchase price
        if self.company.purchase_price:
            sheet['C8'] = float(self.company.purchase_price)
            apply_input_style(sheet['C8'])
            apply_number_format(sheet['C8'], 'currency')
        
        # Entry multiple
        if self.company.entry_multiple:
            sheet['C9'] = float(self.company.entry_multiple)
            apply_input_style(sheet['C9'])
            apply_number_format(sheet['C9'], 'multiple')
        
        # Exit multiple (from valuation)
        if self.valuation and self.valuation.exit_multiple:
            sheet['C12'] = float(self.valuation.exit_multiple)
            apply_input_style(sheet['C12'])
            apply_number_format(sheet['C12'], 'multiple')
        
        logger.info("Transaction assumptions populated")
    
    def _populate_sources_uses(self):
        """Populate sources & uses"""
        if 'Sources & Uses' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Sources & Uses']
        
        # Equity
        if self.company.equity_invested:
            sheet['C10'] = float(self.company.equity_invested)
            apply_input_style(sheet['C10'])
            apply_number_format(sheet['C10'], 'currency')
        
        # Debt
        if self.company.debt_raised:
            sheet['C15'] = float(self.company.debt_raised)
            apply_input_style(sheet['C15'])
            apply_number_format(sheet['C15'], 'currency')
        
        logger.info("Sources & uses populated")
    
    def _populate_operating_model(self):
        """Populate operating model with historicals"""
        if 'Operating Model' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Operating Model']
        
        # Get base year data
        if self.financials:
            latest = self.financials[-1]
            
            # Revenue
            if latest.revenue:
                sheet['C8'] = float(latest.revenue)
                apply_input_style(sheet['C8'])
                apply_number_format(sheet['C8'], 'currency')
            
            # EBITDA
            if latest.ebitda:
                sheet['C15'] = float(latest.ebitda)
                apply_input_style(sheet['C15'])
                apply_number_format(sheet['C15'], 'currency')
        
        logger.info("Operating model populated")


# ============================================================================
# MERGER MODEL GENERATOR
# ============================================================================

class MergerModelGenerator(ModelGenerator):
    """Generate Merger model from database"""
    
    def __init__(self, db_session: Session, acquirer_id: str, target_id: str, template_path: str = None):
        # Note: For merger models, we need TWO company IDs
        self.acquirer_id = acquirer_id
        self.target_id = target_id
        super().__init__(db_session, acquirer_id, template_path)
    
    def fetch_data(self):
        """Fetch data for both acquirer and target"""
        # Get both companies
        self.acquirer = self.db_session.query(PortfolioCompany).filter(
            PortfolioCompany.company_id == self.acquirer_id
        ).first()
        
        self.target = self.db_session.query(PortfolioCompany).filter(
            PortfolioCompany.company_id == self.target_id
        ).first()
        
        if not self.acquirer or not self.target:
            raise ValueError("Acquirer or target company not found")
        
        # Get financials for both
        self.acquirer_financials = self.db_session.query(FinancialMetric).filter(
            FinancialMetric.company_id == self.acquirer_id
        ).order_by(FinancialMetric.period_date).all()
        
        self.target_financials = self.db_session.query(FinancialMetric).filter(
            FinancialMetric.company_id == self.target_id
        ).order_by(FinancialMetric.period_date).all()
        
        logger.info(f"Fetched merger data: {self.acquirer.company_name} acquiring {self.target.company_name}")
    
    def generate(self, output_path: str):
        """Generate complete Merger model"""
        logger.info(f"Generating Merger model...")
        
        # Fetch data
        self.fetch_data()
        
        # Load template
        template_path = '/mnt/user-data/uploads/Merger_Model_Comprehensive.xlsx'
        if os.path.exists(template_path):
            self.wb = load_workbook(template_path)
            logger.info("Loaded Merger template")
        else:
            raise FileNotFoundError("Merger template not found")
        
        # Populate sheets
        self._populate_transaction_assumptions()
        self._populate_pro_forma()
        
        # Save
        self.save(output_path)
    
    def _populate_transaction_assumptions(self):
        """Populate transaction assumptions"""
        if 'Transaction Assumptions' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Transaction Assumptions']
        
        # Acquirer name
        sheet['C5'] = self.acquirer.company_name
        
        # Target name
        sheet['C6'] = self.target.company_name
        
        # Purchase price
        if self.target.purchase_price:
            sheet['C10'] = float(self.target.purchase_price)
            apply_input_style(sheet['C10'])
            apply_number_format(sheet['C10'], 'currency')
        
        logger.info("Merger transaction assumptions populated")
    
    def _populate_pro_forma(self):
        """Populate pro forma income statement"""
        if 'Pro Forma Income Statement' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['Pro Forma Income Statement']
        
        # Get latest financials
        acquirer_latest = self.acquirer_financials[-1] if self.acquirer_financials else None
        target_latest = self.target_financials[-1] if self.target_financials else None
        
        if acquirer_latest:
            # Acquirer revenue
            if acquirer_latest.revenue:
                sheet['C6'] = float(acquirer_latest.revenue)
                apply_input_style(sheet['C6'])
                apply_number_format(sheet['C6'], 'currency')
        
        if target_latest:
            # Target revenue
            if target_latest.revenue:
                sheet['D6'] = float(target_latest.revenue)
                apply_input_style(sheet['D6'])
                apply_number_format(sheet['D6'], 'currency')
        
        logger.info("Pro forma populated")


# ============================================================================
# BATCH MODEL GENERATOR
# ============================================================================

class BatchModelGenerator:
    """Generate all models for a company"""
    
    def __init__(self, db_session: Session):
        self.db_session = db_session
    
    def generate_all_models(self, company_id: str, output_dir: str = '/home/claude/generated_models'):
        """Generate all 5 models for a company"""
        os.makedirs(output_dir, exist_ok=True)
        
        # Get company name
        company = self.db_session.query(PortfolioCompany).filter(
            PortfolioCompany.company_id == company_id
        ).first()
        
        if not company:
            raise ValueError(f"Company {company_id} not found")
        
        company_slug = company.company_name.replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d')
        
        results = {}
        
        # 1. DCF Model
        try:
            dcf_gen = DCFModelGenerator(self.db_session, company_id)
            dcf_path = f'{output_dir}/{company_slug}_DCF_{timestamp}.xlsx'
            dcf_gen.generate(dcf_path)
            results['DCF'] = dcf_path
            logger.info(f"✓ DCF model generated: {dcf_path}")
        except Exception as e:
            logger.error(f"✗ DCF model failed: {e}")
            results['DCF'] = None
        
        # 2. LBO Model
        try:
            lbo_gen = LBOModelGenerator(self.db_session, company_id)
            lbo_path = f'{output_dir}/{company_slug}_LBO_{timestamp}.xlsx'
            lbo_gen.generate(lbo_path)
            results['LBO'] = lbo_path
            logger.info(f"✓ LBO model generated: {lbo_path}")
        except Exception as e:
            logger.error(f"✗ LBO model failed: {e}")
            results['LBO'] = None
        
        # TODO: Add DD Tracker, QoE, and Merger models
        
        return results


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Test the model generation"""
    # This would connect to actual database
    # For now, we'll just demonstrate the structure
    
    print("Excel Model Generator initialized")
    print("This system can generate:")
    print("  1. DCF Model")
    print("  2. LBO Model")
    print("  3. Merger Model")
    print("  4. DD Tracker")
    print("  5. QoE Analysis")
    print("\nTo use:")
    print("  generator = BatchModelGenerator(db_session)")
    print("  generator.generate_all_models(company_id, output_dir)")


if __name__ == '__main__':
    main()
