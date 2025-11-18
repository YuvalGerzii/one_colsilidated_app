"""
Excel Report Generator for Lease Analysis
==========================================

Generates professional, formatted Excel reports with:
- Executive summary dashboard
- Detailed rent roll
- Lease maturity schedule
- Mark-to-market analysis
- Rent growth projections
- Issues and flags
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
from decimal import Decimal

from lease_analyzer import RentRollEntry, RentRollAnalysis


class LeaseReportGenerator:
    """Generate comprehensive Excel reports for lease analysis"""
    
    # Color scheme
    COLORS = {
        'header': '003366',  # Dark blue
        'subheader': '4472C4',  # Medium blue
        'highlight': 'FFC000',  # Orange
        'good': '70AD47',  # Green
        'warning': 'FFC000',  # Orange
        'critical': 'C00000',  # Red
        'light_gray': 'F2F2F2'
    }
    
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
    
    def generate_report(
        self,
        rent_roll: List[RentRollEntry],
        analysis: RentRollAnalysis,
        report_data: Dict,
        output_path: str
    ):
        """
        Generate complete Excel report
        
        Args:
            rent_roll: List of rent roll entries
            analysis: Analysis results
            report_data: Complete report data from service
            output_path: Path to save Excel file
        """
        print(f"\n📊 Generating Excel report: {output_path}")
        
        # Create sheets
        self._create_executive_summary(analysis)
        self._create_detailed_rent_roll(rent_roll)
        self._create_lease_maturity_schedule(report_data['lease_maturity_schedule'])
        self._create_mark_to_market(report_data['mark_to_market_opportunities'])
        self._create_rent_projections(report_data['rent_growth_projections'])
        self._create_issues_and_flags(report_data['issues_and_flags'])
        
        # Save workbook
        self.wb.save(output_path)
        print(f"✅ Report saved: {output_path}")
        
        return output_path
    
    def _create_executive_summary(self, analysis: RentRollAnalysis):
        """Create executive summary dashboard"""
        ws = self.wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = f"Rent Roll Analysis - {analysis.property_name}"
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], 
                                     fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A2'] = f"Analysis Date: {analysis.analysis_date}"
        ws['A2'].font = Font(size=11, italic=True)
        ws.merge_cells('A2:F2')
        
        # Key Metrics Section
        row = 4
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['subheader'], 
                                         end_color=self.COLORS['subheader'], 
                                         fill_type='solid')
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        
        # Occupancy metrics
        metrics = [
            ("Total Square Feet", f"{analysis.total_square_feet:,}", "SF"),
            ("Occupied Square Feet", f"{analysis.occupied_square_feet:,}", "SF"),
            ("Vacant Square Feet", f"{analysis.vacant_square_feet:,}", "SF"),
            ("Physical Occupancy", f"{analysis.physical_occupancy_rate:.1f}", "%"),
            ("Economic Occupancy", f"{analysis.economic_occupancy_rate:.1f}", "%"),
            ("", "", ""),  # Spacer
            ("Number of Tenants", f"{analysis.number_of_tenants}", ""),
            ("Total Annual Rent", f"${analysis.total_annual_rent:,.0f}", ""),
            ("Total Market Rent", f"${analysis.total_market_rent:,.0f}", ""),
            ("Weighted Avg Rent/SF", f"${analysis.weighted_avg_rent_psf:.2f}", "/SF"),
            ("Market Rent/SF", f"${analysis.weighted_avg_market_rent_psf:.2f}", "/SF"),
            ("", "", ""),  # Spacer
            ("Loss to Lease", f"${analysis.total_loss_to_lease:,.0f}", "/year"),
            ("Loss to Lease %", f"{analysis.loss_to_lease_percentage:.1f}", "%"),
            ("", "", ""),  # Spacer
            ("WALT (months)", f"{analysis.weighted_avg_lease_term_months:.1f}", "months"),
            ("12-Month Rollover Risk", f"{analysis.rollover_risk_percentage:.1f}", "%"),
            ("Leases Expiring (12M)", f"{analysis.leases_expiring_12m}", "leases"),
            ("SF Expiring (12M)", f"{analysis.leases_expiring_12m_sf:,}", "SF"),
            ("", "", ""),  # Spacer
            ("Top 5 Concentration", f"{analysis.top_5_tenant_concentration:.1f}", "%"),
            ("Largest Tenant", analysis.largest_tenant_name, ""),
            ("Largest Tenant %", f"{analysis.largest_tenant_percentage:.1f}", "%"),
        ]
        
        for label, value, unit in metrics:
            if label:  # Skip spacers for formatting
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'C{row}'] = unit
                
                # Format label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['light_gray'],
                                                 end_color=self.COLORS['light_gray'],
                                                 fill_type='solid')
                
                # Format value
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                
                # Color code certain metrics
                if 'Occupancy' in label and float(value.replace('%', '').replace(',', '')) < 85:
                    ws[f'B{row}'].font = Font(color=self.COLORS['warning'], bold=True)
                elif 'Loss to Lease' in label and float(value.replace('$', '').replace(',', '').replace('%', '')) > 5:
                    ws[f'B{row}'].font = Font(color=self.COLORS['critical'], bold=True)
                elif 'Rollover Risk' in label and float(value.replace('%', '')) > 30:
                    ws[f'B{row}'].font = Font(color=self.COLORS['warning'], bold=True)
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _create_detailed_rent_roll(self, rent_roll: List[RentRollEntry]):
        """Create detailed rent roll sheet"""
        ws = self.wb.create_sheet("Rent Roll")
        
        # Headers
        headers = [
            'Unit', 'Tenant', 'Status', 'SF', 'Lease Start', 'Lease End',
            'Mo. Rent', 'Annual Rent', 'Rent/SF', 'Market/SF', 'Loss to Lease',
            'Months Left', 'Risk', 'Credit', 'Renewal %', 'Lease Type'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for idx, entry in enumerate(rent_roll, 2):
            ws.cell(idx, 1, entry.unit_number)
            ws.cell(idx, 2, entry.tenant_name)
            ws.cell(idx, 3, entry.status)
            ws.cell(idx, 4, entry.square_feet)
            ws.cell(idx, 5, entry.lease_start or 'N/A')
            ws.cell(idx, 6, entry.lease_end or 'N/A')
            ws.cell(idx, 7, entry.monthly_rent).number_format = '$#,##0'
            ws.cell(idx, 8, entry.annual_rent).number_format = '$#,##0'
            ws.cell(idx, 9, entry.rent_per_sf).number_format = '$#,##0.00'
            ws.cell(idx, 10, entry.market_rent_per_sf).number_format = '$#,##0.00'
            ws.cell(idx, 11, entry.loss_to_lease_annual).number_format = '$#,##0'
            ws.cell(idx, 12, entry.months_remaining or 'N/A')
            ws.cell(idx, 13, entry.expiration_risk)
            ws.cell(idx, 14, entry.credit_rating)
            ws.cell(idx, 15, f"{entry.renewal_probability * 100:.0f}%")
            ws.cell(idx, 16, entry.lease_type)
            
            # Color code risk
            risk_cell = ws.cell(idx, 13)
            if entry.expiration_risk == 'CRITICAL':
                risk_cell.fill = PatternFill(start_color=self.COLORS['critical'],
                                            end_color=self.COLORS['critical'],
                                            fill_type='solid')
                risk_cell.font = Font(color='FFFFFF', bold=True)
            elif entry.expiration_risk == 'HIGH':
                risk_cell.fill = PatternFill(start_color=self.COLORS['warning'],
                                            end_color=self.COLORS['warning'],
                                            fill_type='solid')
            
            # Highlight vacant units
            if entry.status == 'Vacant':
                for col in range(1, 17):
                    ws.cell(idx, col).fill = PatternFill(start_color='FFE699',
                                                         end_color='FFE699',
                                                         fill_type='solid')
        
        # Auto-adjust column widths
        for col in range(1, 17):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Make tenant column wider
        ws.column_dimensions['B'].width = 25
    
    def _create_lease_maturity_schedule(self, maturity_data: List[Dict]):
        """Create lease maturity schedule"""
        ws = self.wb.create_sheet("Lease Maturity")
        
        # Title
        ws['A1'] = "LEASE MATURITY SCHEDULE"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Unit', 'Tenant', 'SF', 'Expiration', 'Months Left', 'Annual Rent', 'Risk']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLORS['light_gray'],
                                   end_color=self.COLORS['light_gray'],
                                   fill_type='solid')
        
        # Data
        for idx, item in enumerate(maturity_data, 4):
            ws.cell(idx, 1, item['Unit'])
            ws.cell(idx, 2, item['Tenant'])
            ws.cell(idx, 3, item['SF'])
            ws.cell(idx, 4, item['Expiration'])
            ws.cell(idx, 5, item['Months Remaining'])
            ws.cell(idx, 6, item['Annual Rent']).number_format = '$#,##0'
            ws.cell(idx, 7, item['Risk'])
            
            # Color code risk
            risk_cell = ws.cell(idx, 7)
            if item['Risk'] == 'CRITICAL':
                risk_cell.fill = PatternFill(start_color=self.COLORS['critical'],
                                            end_color=self.COLORS['critical'],
                                            fill_type='solid')
                risk_cell.font = Font(color='FFFFFF', bold=True)
            elif item['Risk'] == 'HIGH':
                risk_cell.fill = PatternFill(start_color=self.COLORS['warning'],
                                            end_color=self.COLORS['warning'],
                                            fill_type='solid')
        
        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 25
    
    def _create_mark_to_market(self, mtm_data: List[Dict]):
        """Create mark-to-market analysis"""
        ws = self.wb.create_sheet("Mark-to-Market")
        
        # Title
        ws['A1'] = "MARK-TO-MARKET OPPORTUNITIES"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Unit', 'Tenant', 'SF', 'Current $/SF', 'Market $/SF', 
                  'Gap $/SF', 'Annual Opportunity', 'Expiration']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLORS['light_gray'],
                                   end_color=self.COLORS['light_gray'],
                                   fill_type='solid')
        
        # Data
        total_opportunity = 0
        for idx, item in enumerate(mtm_data, 4):
            ws.cell(idx, 1, item['Unit'])
            ws.cell(idx, 2, item['Tenant'])
            ws.cell(idx, 3, item['SF'])
            ws.cell(idx, 4, item['Current Rent/SF']).number_format = '$#,##0.00'
            ws.cell(idx, 5, item['Market Rent/SF']).number_format = '$#,##0.00'
            ws.cell(idx, 6, item['Gap/SF']).number_format = '$#,##0.00'
            ws.cell(idx, 7, item['Annual Opportunity']).number_format = '$#,##0'
            ws.cell(idx, 8, item['Expiration'])
            
            # Highlight opportunity
            opp_cell = ws.cell(idx, 7)
            if item['Annual Opportunity'] > 50000:
                opp_cell.fill = PatternFill(start_color=self.COLORS['highlight'],
                                           end_color=self.COLORS['highlight'],
                                           fill_type='solid')
                opp_cell.font = Font(bold=True)
            
            total_opportunity += item['Annual Opportunity']
        
        # Total row
        total_row = len(mtm_data) + 5
        ws[f'F{total_row}'] = "TOTAL OPPORTUNITY:"
        ws[f'F{total_row}'].font = Font(bold=True)
        ws[f'G{total_row}'] = total_opportunity
        ws[f'G{total_row}'].number_format = '$#,##0'
        ws[f'G{total_row}'].font = Font(bold=True, size=12)
        ws[f'G{total_row}'].fill = PatternFill(start_color=self.COLORS['highlight'],
                                               end_color=self.COLORS['highlight'],
                                               fill_type='solid')
        
        # Auto-adjust columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 25
    
    def _create_rent_projections(self, projection_data: List[Dict]):
        """Create rent growth projections"""
        ws = self.wb.create_sheet("Rent Projections")
        
        # Title
        ws['A1'] = "RENT GROWTH PROJECTIONS (5-YEAR)"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Assumptions
        ws['A3'] = "Assumptions:"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = "• 3% organic rent growth per year"
        ws['A5'] = "• 20% loss-to-lease capture per year"
        
        # Headers
        headers = ['Year', 'Base Growth', 'LTL Capture', 'Total Rent', 'Growth %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(7, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLORS['light_gray'],
                                   end_color=self.COLORS['light_gray'],
                                   fill_type='solid')
        
        # Data
        for idx, item in enumerate(projection_data, 8):
            ws.cell(idx, 1, f"Year {item['Year']}")
            ws.cell(idx, 2, item['Base Rent Growth']).number_format = '$#,##0'
            ws.cell(idx, 3, item['Loss-to-Lease Capture']).number_format = '$#,##0'
            ws.cell(idx, 4, item['Total Projected Rent']).number_format = '$#,##0'
            ws.cell(idx, 5, item['Growth %']).number_format = '0.0%'
            
            # Highlight year 5
            if item['Year'] == 5:
                for col in range(1, 6):
                    cell = ws.cell(idx, col)
                    cell.fill = PatternFill(start_color=self.COLORS['good'],
                                           end_color=self.COLORS['good'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
        
        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_issues_and_flags(self, issues_data: List[Dict]):
        """Create issues and flags summary"""
        ws = self.wb.create_sheet("Issues & Flags")
        
        # Title
        ws['A1'] = "ISSUES & ACTION ITEMS"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Issue Type', 'Severity', 'Count', 'Impact', 'Recommended Action']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLORS['light_gray'],
                                   end_color=self.COLORS['light_gray'],
                                   fill_type='solid')
        
        # Data
        for idx, issue in enumerate(issues_data, 4):
            ws.cell(idx, 1, issue['Type'])
            ws.cell(idx, 2, issue['Severity'])
            ws.cell(idx, 3, issue['Count'])
            ws.cell(idx, 4, issue['Impact'])
            ws.cell(idx, 5, issue['Action'])
            
            # Color code severity
            severity_cell = ws.cell(idx, 2)
            if issue['Severity'] == 'HIGH':
                severity_cell.fill = PatternFill(start_color=self.COLORS['critical'],
                                                end_color=self.COLORS['critical'],
                                                fill_type='solid')
                severity_cell.font = Font(color='FFFFFF', bold=True)
            elif issue['Severity'] == 'MEDIUM':
                severity_cell.fill = PatternFill(start_color=self.COLORS['warning'],
                                                end_color=self.COLORS['warning'],
                                                fill_type='solid')
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 50


def generate_comprehensive_report(
    rent_roll: List[RentRollEntry],
    analysis: RentRollAnalysis,
    report_data: Dict,
    output_path: str
) -> str:
    """
    Generate comprehensive Excel report
    
    Args:
        rent_roll: List of rent roll entries
        analysis: Analysis results
        report_data: Complete report data
        output_path: Path to save Excel file
        
    Returns:
        Path to generated report
    """
    generator = LeaseReportGenerator()
    return generator.generate_report(rent_roll, analysis, report_data, output_path)


if __name__ == "__main__":
    print("Lease Report Generator Module")
    print("Import this module and use generate_comprehensive_report()")
